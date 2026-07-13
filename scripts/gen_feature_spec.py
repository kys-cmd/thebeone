# -*- coding: utf-8 -*-
"""
BE ONE 비원아카데미 기능명세서 생성 스크립트
GitHub thebeone 소스코드 분석 기반
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ---------- 공통 스타일 ----------
TITLE_FONT = Font(name="맑은 고딕", size=16, bold=True, color="FFFFFF")
HEAD_FONT = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
CELL_FONT = Font(name="맑은 고딕", size=10, color="222222")
BOLD_FONT = Font(name="맑은 고딕", size=10, bold=True, color="222222")

PRIMARY = PatternFill("solid", fgColor="4A2FBD")     # 보라 (BE ONE 브랜드)
HEAD_FILL = PatternFill("solid", fgColor="6C4FE0")
SUB_FILL = PatternFill("solid", fgColor="EDE9FB")
ALT_FILL = PatternFill("solid", fgColor="F7F5FE")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

thin = Side(style="thin", color="D9D4EC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def title_block(ws, text, ncols, sub=None):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.font = TITLE_FONT
    c.fill = PRIMARY
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 34
    start = 2
    if sub:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        s = ws.cell(row=2, column=1, value=sub)
        s.font = Font(name="맑은 고딕", size=9, italic=True, color="666666")
        s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[2].height = 18
        start = 3
    return start


def write_table(ws, start_row, headers, rows, widths, cat_col=None):
    # header
    for i, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=i, value=h)
    style_header(ws, start_row, len(headers))
    ws.row_dimensions[start_row].height = 24
    # widths
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # rows
    r = start_row + 1
    prev_cat = None
    for row in rows:
        for i, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=i, value=val)
            cell.font = CELL_FONT
            cell.alignment = LEFT
            cell.border = BORDER
            if (r % 2) == 0:
                cell.fill = ALT_FILL
            else:
                cell.fill = WHITE_FILL
        # emphasise first column (category)
        if cat_col is not None:
            ws.cell(row=r, column=cat_col).font = BOLD_FONT
        r += 1
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    return r


# =====================================================================
# 1. 개요 (Overview)
# =====================================================================
ws = wb.active
ws.title = "1.개요"
ws.sheet_view.showGridLines = False
start = title_block(ws, "BE ONE 비원아카데미 · 기능명세서",
                    4, sub="정규강의 · 특강 · 커뮤니티 · 라이브가 결합된 온라인 교육 플랫폼  |  작성일 2026-07-13")

overview = [
    ("프로젝트명", "BE ONE (비원아카데미)", "카테고리", "온라인 교육 / 커뮤니티 플랫폼"),
    ("서비스 개요", "정규강의, 온·오프라인 특강, 회원 전용 강의, 라이브 방송과 커뮤니티/실시간 채팅을 결합한 프리미엄 교육 플랫폼", "", ""),
    ("Frontend", "React 19, TypeScript, Vite 6, React Router 7, Zustand, TailwindCSS 4, shadcn/ui, Recharts, TipTap 에디터", "", ""),
    ("Backend", "Node.js(Express) SSR 서버(server.ts), Netlify Functions", "", ""),
    ("Database / Infra", "Supabase(PostgreSQL, Auth, Realtime, Storage), Docker, Cloud Build / Netlify 배포", "", ""),
    ("결제(PG)", "이니시스(KG Inicis) 통합 결제 연동, 마일리지 차감 결제 지원", "", ""),
    ("인증", "이메일/비밀번호, Google OAuth, 이메일 OTP 기반 비밀번호 재설정", "", ""),
    ("영상", "Vimeo 임베드 + Vimeo Player SDK(수강 진도율 추적)", "", ""),
    ("주요 사용자 권한", "super_admin / admin / beone_member / paid_member / regular_member / user", "", ""),
]
r = start
for label, val, label2, val2 in overview:
    ws.cell(row=r, column=1, value=label).font = BOLD_FONT
    ws.cell(row=r, column=1).fill = SUB_FILL
    if label2:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=2)
        ws.cell(row=r, column=2, value=val)
        ws.cell(row=r, column=3, value=label2).font = BOLD_FONT
        ws.cell(row=r, column=3).fill = SUB_FILL
        ws.cell(row=r, column=4, value=val2)
    else:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        ws.cell(row=r, column=2, value=val)
    for c in range(1, 5):
        cell = ws.cell(row=r, column=c)
        cell.font = BOLD_FONT if c in (1, 3) else CELL_FONT
        cell.alignment = WRAP
        cell.border = BORDER
    ws.row_dimensions[r].height = 30
    r += 1

for col, w in zip("ABCD", [18, 40, 16, 40]):
    ws.column_dimensions[col].width = w

# 시트 안내
r += 1
ws.cell(row=r, column=1, value="■ 시트 구성 안내").font = Font(name="맑은 고딕", size=11, bold=True, color="4A2FBD")
r += 1
guide = [
    "2.기능명세  : 대분류/중분류별 전체 기능 목록 및 상세 설명 (핵심)",
    "3.사용자화면 : 사용자(수강생) 화면 및 라우트 정의",
    "4.관리자화면 : 관리자(Admin) 화면 및 관리 기능 정의",
    "5.데이터모델 : 주요 DB 테이블(엔티티) 정의",
    "6.API명세   : 서버 REST API 및 core-api 액션 목록",
    "7.권한정책  : 사용자 등급(Role) 및 접근 권한 정책",
]
for g in guide:
    ws.cell(row=r, column=1, value=g).font = CELL_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.cell(row=r, column=1).alignment = LEFT
    r += 1


# =====================================================================
# 2. 기능명세 (핵심)
# =====================================================================
ws = wb.create_sheet("2.기능명세")
ws.sheet_view.showGridLines = False
start = title_block(ws, "기능명세 (Functional Specification)", 6)
headers = ["ID", "대분류", "중분류", "기능명", "기능 설명", "권한"]
widths = [8, 14, 18, 24, 62, 14]

# (대분류, 중분류, 기능명, 설명, 권한)
feat = [
    # 회원/인증
    ("회원/인증", "회원가입", "이메일 회원가입", "이메일·비밀번호로 가입, 이름·닉네임·연락처·성별·생년월일 입력", "비회원"),
    ("회원/인증", "회원가입", "소셜 로그인", "Google OAuth 로그인/가입, 미입력 추가정보는 마이페이지로 유도", "비회원"),
    ("회원/인증", "로그인", "로그인", "이메일/비밀번호 및 소셜 로그인, 세션 자동 복원", "비회원"),
    ("회원/인증", "로그인", "이메일 찾기", "이름·연락처로 가입 이메일 조회", "비회원"),
    ("회원/인증", "비밀번호", "비밀번호 재설정", "이메일 OTP 발송·검증 후 비밀번호 변경", "비회원"),
    ("회원/인증", "세션", "권한/세션 관리", "Zustand 스토어로 로그인 상태·권한 유지, 관리자 라우트 보호", "전체"),

    # 강의 (수강)
    ("강의", "강의 목록", "전체 강의 조회", "발행된 강의 목록 조회, 카테고리(정규/특강/라이브)별 필터", "전체"),
    ("강의", "강의 목록", "정규 강의", "정규강의 목록 및 상세 노출", "전체"),
    ("강의", "강의 목록", "온·오프라인 특강", "온라인/오프라인 특강 분리 노출, 정원·수강인원 표시", "전체"),
    ("강의", "강의 목록", "회원 전용 강의(BE ONE)", "회원 등급(min_member_grade) 이상만 접근 가능한 전용 강의", "회원등급"),
    ("강의", "강의 상세", "강의 상세 페이지", "커리큘럼·강사정보·소개·수강후기·수강신청 버튼 노출", "전체"),
    ("강의", "강의 상세", "동적 레이아웃 렌더링", "블록(영상/이미지/텍스트/제목) 기반 커스텀 상세 구성", "전체"),
    ("강의", "수강신청", "수강 등록", "결제 완료 또는 무료/수동 등록으로 수강권 부여, 만료일 관리", "회원"),
    ("강의", "수강신청", "기간제 수강", "duration_days 기반 수강 만료일 자동 계산", "회원"),
    ("강의", "학습", "학습 플레이어", "Vimeo 영상 재생, 커리큘럼 이동, 학습 자료 다운로드", "수강생"),
    ("강의", "학습", "진도율 추적", "Vimeo SDK로 영상 시청 진도 저장·복원", "수강생"),
    ("강의", "학습", "레슨 접근 검증", "서버에서 수강권·무료여부 확인 후 레슨 접근 허용", "수강생"),

    # 강사
    ("강사", "강사", "강사 정보", "강사 프로필(이름·직함·소개·이미지·태그) 제공 및 강의 연결", "전체"),

    # 수강후기
    ("수강후기", "후기", "후기 작성/수정/삭제", "수강생이 강의별 별점·후기 작성 및 본인 후기 관리", "수강생"),
    ("수강후기", "후기", "후기 조회", "강의 상세·홈에서 후기 및 베스트 후기 노출", "전체"),

    # 결제/주문
    ("결제/주문", "장바구니", "장바구니", "강의 담기, 선택/일괄 삭제, 선택 결제", "회원"),
    ("결제/주문", "결제", "이니시스 결제", "KG이니시스 PG 통합 결제(카드 등), 결제창 호출·서명", "회원"),
    ("결제/주문", "결제", "마일리지 결제", "보유 마일리지 차감 후 잔액 결제", "회원"),
    ("결제/주문", "결제", "결제 콜백/검증", "결제 결과 수신·검증, 주문 상태(PAID) 처리 및 수강권 부여", "시스템"),
    ("결제/주문", "주문", "주문 내역 조회", "마이페이지에서 결제/주문 이력 확인", "회원"),
    ("결제/주문", "주문", "결제 실패 알림", "결제 실패 시 관리자/사용자 알림 발송", "시스템"),

    # 커뮤니티
    ("커뮤니티", "게시판", "게시글 CRUD", "커뮤니티 게시글 작성·수정·삭제(에디터, 이미지/파일/링크 첨부)", "회원"),
    ("커뮤니티", "게시판", "게시글 유형", "공지/일반/미션템플릿/미션인증 등 유형별 게시", "회원/관리자"),
    ("커뮤니티", "게시판", "좋아요/댓글", "게시글 좋아요, 댓글 작성·수정·삭제", "회원"),
    ("커뮤니티", "게시판", "이모지 반응", "게시글·메시지 이모지 반응 토글", "회원"),
    ("커뮤니티", "게시판", "출석/투표/할일", "미션 게시글 출석 체크, 투표, 할 일(To-do) 체크", "회원"),
    ("커뮤니티", "게시판", "게시글 고정", "관리자/작성자 게시글 상단 고정", "관리자"),
    ("커뮤니티", "커뮤니티", "커뮤니티 가입", "결제 완료 시 강의 커뮤니티 자동 가입, 초대/탈퇴", "회원"),
    ("커뮤니티", "커뮤니티", "시즌/게시판 유형", "시즌제·강의연동·일반 게시판 유형 관리", "관리자"),

    # 채팅
    ("실시간 채팅", "채팅", "실시간 메시지", "Supabase Realtime 기반 커뮤니티/방 채팅", "회원"),
    ("실시간 채팅", "채팅", "채팅방 관리", "채팅방 개설·입장·퇴장·삭제, 공개/비공개방", "회원"),
    ("실시간 채팅", "채팅", "1:1 채팅", "회원 검색 후 1:1 대화방 생성", "회원"),
    ("실시간 채팅", "채팅", "미디어 전송", "이미지/파일/GIF/스티커/이모지 메시지 전송", "회원"),
    ("실시간 채팅", "채팅", "안읽음/읽음", "미읽음 카운트 집계 및 읽음 처리", "회원"),
    ("실시간 채팅", "채팅", "플로팅/팝업 채팅", "전 화면 플로팅 채팅 위젯 및 팝업창 채팅", "회원"),

    # 라이브
    ("라이브", "라이브", "라이브 방송", "Vimeo 임베드 실시간 방송 시청, 준비중 상태 표시", "전체/회원"),
    ("라이브", "라이브", "라이브 채팅", "방송 중 실시간 채팅", "회원"),

    # 마이페이지
    ("마이페이지", "내정보", "프로필 관리", "이름·닉네임·연락처·성별·생년월일 등 회원정보 수정", "회원"),
    ("마이페이지", "내정보", "비밀번호 변경", "로그인 상태에서 비밀번호 변경", "회원"),
    ("마이페이지", "수강", "내 강의", "수강 중/신청한 강의 목록 및 학습 이동", "회원"),
    ("마이페이지", "결제", "결제 내역/마일리지", "결제 이력 및 보유 마일리지 조회", "회원"),

    # 알림
    ("알림", "알림", "알림 수신", "게시글/문의답변/결제 등 알림 생성·조회·읽음", "회원"),

    # 고객지원/정적
    ("고객지원", "지원", "공지사항", "공지 목록/상세 조회", "전체"),
    ("고객지원", "지원", "1:1 문의", "문의 등록 및 관리자 답변 확인", "전체/회원"),
    ("고객지원", "지원", "FAQ", "자주 묻는 질문 조회", "전체"),
    ("고객지원", "지원", "건의하기", "서비스 개선 건의 접수", "전체"),
    ("고객지원", "지원", "마스터 지원", "강사/마스터 지원 신청", "전체"),
    ("정적/소개", "소개", "회사/서비스 소개", "About, 라이프코칭, 테크트리 등 브랜드 소개 페이지", "전체"),
    ("정적/소개", "약관", "약관/개인정보", "이용약관·개인정보처리방침", "전체"),

    # 관리자 (요약, 상세는 4번 시트)
    ("관리자", "대시보드", "관리자 대시보드", "오늘 매출·가입자·결제 건수·주간 매출 차트 등 지표", "관리자"),
    ("관리자", "강의관리", "강의 관리/편집", "강의 등록·수정·삭제, 커리큘럼·레이아웃 편집", "관리자"),
    ("관리자", "회원관리", "회원/수강생 관리", "회원 조회·권한변경·비밀번호 재설정·수강권 부여/회수", "관리자"),
    ("관리자", "커뮤니티관리", "커뮤니티 관리", "커뮤니티·게시판 생성/수정/멤버 관리", "관리자"),
    ("관리자", "매출관리", "매출/주문 관리", "주문 조회·환불·재처리·통계", "관리자"),
    ("관리자", "CMS", "디자인/CMS", "메인 배너·HERO·사이트 설정·결제 설정 관리", "관리자"),
    ("관리자", "운영", "후기/문의/오류로그", "후기·1:1문의·클라이언트/결제 오류 로그 관리", "관리자"),
]

rows = [(f"F-{i:03d}", a, b, c, d, e) for i, (a, b, c, d, e) in enumerate(feat, start=1)]
r = write_table(ws, start, headers, rows, widths, cat_col=2)


# =====================================================================
# 3. 사용자 화면
# =====================================================================
ws = wb.create_sheet("3.사용자화면")
ws.sheet_view.showGridLines = False
start = title_block(ws, "사용자(수강생) 화면 정의", 4)
headers = ["화면명", "경로(Route)", "주요 기능", "권한"]
widths = [24, 26, 60, 14]
screens = [
    ("홈", "/", "메인 배너, 추천/카테고리별 강의, 베스트 후기", "전체"),
    ("전체 강의", "/courses", "전체 강의 목록·카테고리 필터", "전체"),
    ("특강", "/special", "온·오프라인 특강 목록", "전체"),
    ("회원 전용", "/beone", "회원 등급 전용 강의", "회원등급"),
    ("라이브", "/live, /live/:id", "라이브 방송 시청·채팅", "전체/회원"),
    ("테크트리", "/techtree", "성장 로드맵 소개", "전체"),
    ("라이프코칭", "/lifecoaching", "라이프코칭 진단·소개", "전체"),
    ("강의 상세", "/course/:id", "커리큘럼·강사·후기·수강신청", "전체"),
    ("학습 플레이어", "/course/:id/learn", "영상 학습·진도율·자료", "수강생"),
    ("커뮤니티", "/community", "게시판·게시글·댓글·반응", "회원"),
    ("채팅", "/chat", "실시간 채팅(방 목록·1:1)", "회원"),
    ("마이페이지", "/mypage", "내 정보·내 강의·결제·마일리지", "회원"),
    ("장바구니", "/cart", "담은 강의·결제", "회원"),
    ("소개", "/about", "브랜드/서비스 소개", "전체"),
    ("로그인/회원가입", "/auth/login, /auth/register", "인증", "비회원"),
    ("비밀번호 재설정", "/auth/reset-password", "OTP 비밀번호 재설정", "비회원"),
    ("소셜 콜백", "/auth/callback", "OAuth 콜백 처리", "시스템"),
    ("결제 콜백", "/payment/callback", "PG 결제 결과 처리", "시스템"),
    ("고객지원", "/support/notice, /faq, /inquiry, /suggestion, /apply", "공지·FAQ·문의·건의·마스터지원", "전체"),
    ("약관", "/legal/terms, /legal/privacy", "이용약관·개인정보처리방침", "전체"),
]
write_table(ws, start, headers, screens, widths, cat_col=1)


# =====================================================================
# 4. 관리자 화면
# =====================================================================
ws = wb.create_sheet("4.관리자화면")
ws.sheet_view.showGridLines = False
start = title_block(ws, "관리자(Admin) 화면 정의", 3, sub="접근 권한: super_admin / admin 전용 (AdminPrivateRoute 보호)")
headers = ["관리 메뉴", "경로(Route)", "주요 기능"]
widths = [22, 30, 66]
admin = [
    ("대시보드", "/admin", "오늘 매출·전체 가입자·결제 건수·주간 매출 차트, 최근 수강신청"),
    ("강의 관리", "/admin/courses", "강의 목록·통계·삭제, 발행 상태 관리"),
    ("강의 편집", "/admin/courses/edit/:id", "강의 정보·커리큘럼·동적 레이아웃·수강기간 편집"),
    ("강사 관리", "/admin/instructors", "강사 등록·수정·삭제, 태그 관리"),
    ("수강 후기 관리", "/admin/reviews", "후기 조회·삭제·베스트 후기 지정"),
    ("디자인/CMS", "/admin/cms", "HERO 디자인·메인 배너·커뮤니티 카테고리 관리"),
    ("회원 관리", "/admin/users", "회원 조회·권한변경·비밀번호 재설정·계정 삭제·재설정 요청 처리"),
    ("수강생 관리", "/admin/students", "수강생·출석 조회, 엑셀 내보내기, 강의별 필터"),
    ("커뮤니티 관리", "/admin/community", "커뮤니티·게시판 생성/수정, 멤버·방 관리, 강의연동 동기화"),
    ("커뮤니티 편집", "/admin/community/create, /edit/:id", "커뮤니티 게시글/공지 작성·편집"),
    ("매출 관리", "/admin/sales", "주문·결제 조회, 환불·재처리, 결제 상세(TID)"),
    ("통계", "/admin/stats", "월별·주별 매출 분석 차트"),
    ("시스템 설정", "/admin/settings", "사이트 설정·결제(PG) 설정 저장"),
    ("고객지원 관리", "/admin/support", "1:1 문의 답변, 공지/FAQ/일정 콘텐츠 관리"),
    ("오류 로그", "/admin/errors", "클라이언트/결제 API 오류 로그 조회·삭제"),
]
write_table(ws, start, headers, admin, widths, cat_col=1)


# =====================================================================
# 5. 데이터 모델
# =====================================================================
ws = wb.create_sheet("5.데이터모델")
ws.sheet_view.showGridLines = False
start = title_block(ws, "데이터 모델 (주요 엔티티)", 3, sub="Supabase PostgreSQL · 금액 BIGINT 정수 관리 · is_deleted Soft Delete 원칙")
headers = ["테이블(엔티티)", "설명", "주요 컬럼"]
widths = [24, 34, 70]
tables = [
    ("profiles", "사용자 프로필/회원정보", "id, email, name, nickname, role, gender, birthdate, mobile_phone, address, mileage, is_deleted"),
    ("courses", "강의", "id, title, price, discount_price, category, status, curriculum(JSONB), layout(JSONB), instructor_id, access_type, min_member_grade, duration_days, is_deleted"),
    ("instructors", "강사", "id, name, title, bio, image, tags"),
    ("enrollments", "수강권(수강 등록)", "user_id, course_id, created_at, expires_at"),
    ("orders", "주문/결제", "id, user_id, course_id, amount, mileage_used, status(pending/PAID/cancelled), payment_tid, created_at"),
    ("reviews", "수강 후기", "id, course_id, user_id, rating, content, is_best, created_at"),
    ("communities", "커뮤니티/게시판", "id, course_id, name, type(season/course/board), post_permission, season_number, is_deleted"),
    ("community_members", "커뮤니티 멤버십", "id, community_id, user_id, joined_at"),
    ("posts", "게시글", "id, community_id, user_id, title, content, type, hashtags, image_urls, file_urls, views, is_pinned, is_deleted"),
    ("comments", "댓글", "id, post_id, user_id, content, created_at"),
    ("messages", "채팅 메시지", "id, community_id, room_id, user_id, content, type(text/image/file/gif/sticker), media_url, reply_to_id"),
    ("chat_rooms", "채팅방", "id, community_id, name, room_type, is_private, members"),
    ("notifications", "알림", "id, user_id, type, content, is_read, created_at"),
    ("banners", "CMS 배너", "id, type(main/course/special/beone_exclusive), image, link, order"),
    ("site_config / payment_settings", "사이트·결제 설정", "site config(JSONB), PG(이니시스) 설정값"),
    ("support/inquiries", "고객지원 콘텐츠", "공지·FAQ·문의·건의, 답변 내용"),
    ("payment_api_logs / error_logs", "결제·오류 로그", "요청/응답 페이로드, 오류 메시지, 발생일시"),
]
write_table(ws, start, headers, tables, widths, cat_col=1)


# =====================================================================
# 6. API 명세
# =====================================================================
ws = wb.create_sheet("6.API명세")
ws.sheet_view.showGridLines = False
start = title_block(ws, "API 명세 (server.ts)", 3, sub="REST 엔드포인트 + /api/core-api 액션 디스패처")
headers = ["구분 / Method", "엔드포인트 · 액션", "설명"]
widths = [22, 40, 62]
apis = [
    ("REST · POST", "/api/init-pay", "이니시스 결제 초기화(결제창 파라미터 발급)"),
    ("REST · POST", "/api/inicis-return", "이니시스 결제 결과 리턴 처리"),
    ("REST · POST", "/api/payment/inicis/sign", "이니시스 결제 서명(sign) 생성"),
    ("REST · POST", "/api/payment/inicis/notify", "이니시스 결제 통지(Notify) 수신"),
    ("REST · POST", "/api/payment/inicis/callback", "이니시스 결제 콜백 처리·검증"),
    ("REST · POST", "/api/auth/reset-request", "비밀번호 재설정 요청 접수"),
    ("REST · POST", "/api/auth/find-email", "가입 이메일 찾기"),
    ("REST · POST", "/api/auth/send-reset-otp", "비밀번호 재설정 OTP 발송"),
    ("REST · POST", "/api/auth/verify-reset-otp", "OTP 검증"),
    ("REST · POST", "/api/auth/confirm-reset-password", "OTP 확인 후 비밀번호 변경"),
    ("REST · POST", "/api/admin/reset-password", "관리자 비밀번호 강제 재설정"),
    ("REST · POST", "/api/admin/delete-auth-user", "인증 사용자 삭제"),
    ("REST · GET", "/api/admin/orphaned-auth-users", "고아 인증계정 조회"),
    ("REST · POST", "/api/admin/bulk-delete-auth-users", "인증계정 일괄 삭제"),
    ("REST · POST", "/api/admin/sync-all-communities", "전체 커뮤니티 멤버 동기화"),
    ("REST · POST", "/api/video/progress", "영상 시청 진도 저장"),
    ("REST · GET", "/api/naver-blog / /api/link-preview", "네이버 블로그 RSS·링크 미리보기"),
    ("REST · GET", "/api/assets/:bucket/*", "Supabase Storage 자산 프록시"),
    ("core-api", "verify-lesson-access", "레슨 접근 권한 검증(수강권/무료/관리자)"),
    ("core-api", "process-payment", "결제 처리·주문 PAID 전환·수강권/커뮤니티 부여·마일리지 차감"),
    ("core-api", "grant-enrollment / revoke-enrollment", "수강권 수동 부여/회수"),
    ("core-api", "join-community / leave-community", "커뮤니티 가입/탈퇴"),
    ("core-api", "sync-communities / sync-all-communities", "커뮤니티 멤버 동기화"),
    ("core-api", "submit-inquiry", "1:1 문의 등록"),
    ("core-api", "admin-modify-role", "회원 권한 변경"),
    ("core-api", "admin-reset-password / admin-delete-auth-user", "회원 비밀번호 재설정·계정 삭제"),
    ("core-api", "admin-delete-review / admin-toggle-best-review", "후기 삭제·베스트 지정"),
    ("core-api", "reprocess-order / refund-order / admin-delete-order", "주문 재처리·환불·삭제"),
    ("core-api", "admin-create-test-order / admin-simulate-payment", "테스트 주문·결제 시뮬레이션"),
    ("core-api", "admin-get/clear-payment-api-logs", "결제 API 로그 조회·삭제"),
    ("core-api", "send-payment-failure-alert", "결제 실패 알림 발송"),
]
write_table(ws, start, headers, apis, widths, cat_col=1)


# =====================================================================
# 7. 권한 정책
# =====================================================================
ws = wb.create_sheet("7.권한정책")
ws.sheet_view.showGridLines = False
start = title_block(ws, "사용자 등급 및 권한 정책", 3)
headers = ["Role (등급)", "설명", "주요 권한"]
widths = [22, 34, 66]
roles = [
    ("super_admin", "최고 관리자", "전체 관리자 기능·시스템 설정·회원/결제 관리 전권"),
    ("admin", "관리자", "강의·회원·커뮤니티·매출·CMS 등 운영 관리"),
    ("beone_member / bione_member", "비원(BE ONE) 회원", "회원 전용 강의·커뮤니티 접근(min_member_grade 충족)"),
    ("paid_member", "유료 회원", "결제 강의 수강·커뮤니티 참여"),
    ("regular_member", "일반 회원", "기본 강의 수강·커뮤니티 참여"),
    ("user", "가입 사용자", "로그인·프로필·장바구니·문의 등 기본 기능"),
    ("비회원(Guest)", "비로그인 방문자", "홈·강의 목록/상세·소개·공지/FAQ 열람"),
]
r = write_table(ws, start, headers, roles, widths, cat_col=1)
r += 1
note = [
    "※ 강의 접근: access_type(automatic/manual) 및 min_member_grade, 수강권(enrollment) 만료일로 제어",
    "※ 커뮤니티 가입: 결제 완료(status=PAID) 시점에 강의 연동 커뮤니티 자동 가입 트리거",
    "※ 관리자 라우트(/admin/*)는 AdminPrivateRoute에서 role 검증 후 접근 허용",
    "※ 회계 원칙: 금액은 BIGINT 정수 관리, 데이터는 is_deleted Soft Delete, 주문 시 결제 상태·금액 엄격 기록",
]
for n in note:
    ws.cell(row=r, column=1, value=n).font = Font(name="맑은 고딕", size=9, color="4A2FBD")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.cell(row=r, column=1).alignment = LEFT
    ws.row_dimensions[r].height = 18
    r += 1


out = "/home/user/thebeone/비원아카데미_기능명세서.xlsx"
wb.save(out)
print("saved:", out)
print("sheets:", wb.sheetnames)
